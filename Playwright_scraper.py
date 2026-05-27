import pandas as pd
from playwright.sync_api import sync_playwright
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

arquivo="books.xlsx"
with pd.ExcelWriter(arquivo,engine="openpyxl") as writer:
    with sync_playwright() as p:
        browser=p.chromium.launch(headless=False)
        pagina=browser.new_page()
        for num_pagina in range(1,51):
            produtos_pagina=[]
            if num_pagina == 1:
                url="https://books.toscrape.com/"
            else:
                url=f"https://books.toscrape.com/catalogue/page-{num_pagina}.html"
            print(f"Coletando página {num_pagina}...")
            pagina.goto(url)
            pagina.wait_for_selector("article.product_pod",timeout=15000)
            livros=pagina.locator("article.product_pod").all()
            for lv in livros:
                nome=lv.locator("h3 a").get_attribute("title")
                preco_texto=lv.locator("p.price_color").inner_text()
                preco=float(preco_texto.replace("£", ""))
                avaliacao=lv.locator("p.star-rating").get_attribute("class")
                avl=avaliacao.replace("star-rating ","")
                dados={
                    "Name":nome,
                    "Price":preco,
                    "Rating":avl
                }
                produtos_pagina.append(dados)
            if produtos_pagina:
                df=pd.DataFrame(produtos_pagina)
                nome_aba=f"Page {num_pagina}"
                df.to_excel(writer,sheet_name=nome_aba,index=False)
                ws=writer.sheets[nome_aba]
                fonte="Segoe UI"
                header_font=Font(name=fonte,size=11,bold=True,color="FFFFFF")
                header_fill = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid")
                zebra_fill = PatternFill(start_color="F2F5F8", end_color="F2F5F8", fill_type="solid")
                borda_fina = Border(
                    left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                    top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9')
                )
                ws.row_dimensions[1].height=26
                for cell in ws[1]:
                    cell.font=header_font
                    cell.fill=header_fill
                    cell.alignment=Alignment(horizontal="center",vertical="center")
                    cell.border=borda_fina
                for row_idx in range(2,ws.max_row+1):
                    ws.row_dimensions[row_idx].height=20
                    fill_atual=zebra_fill if row_idx % 2==0 else PatternFill(fill_type=None)
                    for col_idx in range(1,4):
                        cell=ws.cell(row=row_idx,column=col_idx)
                        cell.font=Font(name=fonte,size=10)
                        cell.fill=fill_atual
                        cell.border=borda_fina
                        if col_idx==1:
                            cell.alignment=Alignment(horizontal='left',vertical='center')
                        elif col_idx==2:
                            cell.alignment=Alignment(horizontal='right',vertical='center')
                            cell.number_format='"£"#,##0.00'
                        else:
                            cell.alignment=Alignment(horizontal='center',vertical='center')
                for col in ws.columns:
                    max_len=max(len(str(cell.value or ''))for cell in col)
                    col_letter=col[0].column_letter
                    ws.column_dimensions[col_letter].width=max(max_len+4,12)
print(f"\nSucesso! O arquivo foi gerado e os livros foram divididos por abas.")
print(f"Verifique o arquivo criado: '{arquivo}'")
